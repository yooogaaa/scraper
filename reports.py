 # 此脚本用于在输入标题/标签关键字和页码范围的条件下，自动化提取199IT、贝恩的报告发布时间、标签、标题和链接
 # ---2024.1.23 update---
 # 1 增加贝恩网页研报资源；2 增加199it的报告标签提取功能
 # 存在问题：1 若一个关键词在标题和标签中均出现，会重复记录到Excel中；2 即使一篇研报有多个标签，Excel中的标签只出现一个值
 # ---2024.1.25 update---
 # 已修复问题1：通过在循环之前创建集合，确保了在整个循环过程中都使用相同的集合来跟踪已经写入的组合，从而防止了重复写入的问题；
 # 已修复问题2：需要将for循环中的多个关键词转换为列表，再使用join转换为一行字符串，再写入Excel中
 # 待优化点：1.爬虫速度慢；2.可以把页码范围改成报告时间
 # ---2024.1.27 update---
 # 1.将用页码搜索改为用年份搜索 2.暂时隐藏贝恩的数据 3.将原先的一个主函数进行拆分，爬取时间优化5秒左右
 # 待优化点：1.确定代码范围的函数改为多线程 2.确定代码范围的函数展现进度
 # ---2024.1.29 update---
 # 1.增加贝恩的数据 2.多线程并未加速，反而拖慢 3.『正在提取报告』会打印两遍 4.若先确定代码范围再展现进度，则严重拖慢运行速度，遂放弃
 # ---2024.1.29 update---
 # 1.已改变打印语句以显示大概的进度 2.输出Excel中增加报告时间列，显示发布日期 3.已尽可能使用正则表达式，2023作为起始年份的运行速度缩短到37秒

import time
import openpyxl as op
import requests
from bs4 import BeautifulSoup
import threading

#记录代码开始的时间
# start_time = time.time()

def scrape_reports_bain(keywords, start_year, headers, worksheet, written_combinations, count):
    num = 0
    # 打印进度
    print("即将完成...")
    while True:
        response_bain = requests.get(f"https://www.bain.cn/news.php?id=15&page={num}", headers = headers)
        html_bain = response_bain.text
        soup_bain = BeautifulSoup(html_bain, 'html.parser')
        # 提取报告的年份信息，输出结果为可迭代的列表
        all_time_bain = soup_bain.find_all('div', class_='card-footer')
        for time_bain in all_time_bain:
            year_string = time_bain.string.strip()[:4]
            if start_year <= year_string:
                report_name_bain = time_bain.find_previous('h5').string.strip()
                report_link_bain = "https://www.bain.cn/" +time_bain.find_previous('div', class_ = 'card-content').a.get('href')
                report_desc_bain = time_bain.find_previous('div', class_ = 'card__desc').string
                report_tag_bain = time_bain.find_previous('span', class_ = 'card__tag').string.strip()
                report_date_bain = time_bain.string.strip()[:10]
                # 生成一个组合文本
                combined_text_bain = f"{report_name_bain}{report_tag_bain}{report_desc_bain}"
                # 若报告标题、报告标签、报告描述中出现关键字，则提取报告
                if any(keyword in combined_text_bain for keyword in keywords):
                    if combined_text_bain not in written_combinations:
                        worksheet.cell(row=count[0], column=3, value=report_name_bain)
                        worksheet.cell(row=count[0], column=4, value=report_link_bain)
                        worksheet.cell(row=count[0], column=2, value=report_tag_bain)
                        worksheet.cell(row=count[0], column=1, value=report_date_bain)
                        count[0] += 1
                        written_combinations.add(combined_text_bain)
        # 设置循环退出条件
        if start_year > year_string:
            break
        num = num + 1

def scrape_reports_199it(keywords, start_year, headers, worksheet, written_combinations, count):
    num = 0
    # 打印进度
    print("正在提取报告...")
    # 循环爬取每一页的报告，直到报告年份小于用户输入的起始年份停止
    while True:
        response_199it = requests.get(f"https://www.199it.com/archives/category/report/page/{num}", headers=headers)
        # 获取并解析199it网页的HTML
        html_199it = response_199it.text
        soup_199it = BeautifulSoup(html_199it, 'html.parser')
        # 提取报告标签和时间列表
        all_tag_199it = soup_199it.find_all('ul', class_="post-categories")
        all_time_199it = soup_199it.find_all('time', class_='entry-date')
        # 循环比较每个报告发布的年份与用户输入的起始年份
        for time_199it, tag_199it in zip(all_time_199it, all_tag_199it):
            year_string = time_199it['datetime'][:4]
            if start_year <= year_string:
                report_name_199it = time_199it.find_previous('h2', class_='entry-title').a.get('title')
                report_link_199it = time_199it.find_previous('h2', class_='entry-title').a.get('href')[2:] # 去掉链接开头的 //。若不确定是否每行都有 //，可以使用replace方法，在这里即report_link = report_link.replace("//", "", 1)
                report_date_199it = time_199it['datetime'][:10]
                tag_199it_helper = tag_199it('a', href=True, class_=False, title=False, string=True)
                # 从HTML标签列表中提取标签字符串，并生成列表，再将列表转化为一行字符串
                report_tag_199it_string = ','.join(map(str, [i.string for i in tag_199it_helper]))
                # 创建一个包含研报标题和标签的字符串集合
                combined_text_199it = f"{report_name_199it}{report_tag_199it_string}"
                # 如果标题或标签包含任一关键字，则写入数据
                if any(keyword in combined_text_199it for keyword in keywords): #因为要查询含有多个关键词的列表，需使用"any"的函数结构
                    # 检查组合是否已经写入，如果没有，则写入并将组合添加到集合中
                    if combined_text_199it not in written_combinations:
                        worksheet.cell(row=count[0], column=3, value=report_name_199it)
                        worksheet.cell(row=count[0], column=4, value=report_link_199it)
                        worksheet.cell(row=count[0], column=2, value=report_tag_199it_string)
                        worksheet.cell(row=count[0], column=1, value=report_date_199it)
                        count[0] += 1
                        written_combinations.add(combined_text_199it)
        # 设置循环退出条件
        if start_year > year_string:
            break
        # 完成上述步骤后，换第二页重复操作
        num = num + 1

def main():
    # 接收用户输入的关键字，使用逗号分隔
    search_keywords = input("请输入关键字，用中文逗号分隔: ").split('，')
    start_year = input("请输入想要查询的起始年份（输入2023，则提取2023至今的报告）： ")
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    # 创建工作簿
    workbook = op.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=3, value="报告名称")
    worksheet.cell(row=1, column=4, value="报告链接")
    # bain增加报告标签和报告时间列
    worksheet.cell(row=1, column=2, value="报告标签")
    worksheet.cell(row=1, column=1, value="发布时间")
    # count是一个计数器，用于记录在 Excel 表格中的行数。初始值为 2，每提取一次报告信息，count每增加 1，这样确保每条报告信息都被写入 Excel 表格的不同行。这是为了防止新的报告覆盖之前的报告数据
    count = [2]
    # 用于跟踪已写入的组合的集合
    written_combinations = set()
    # 调用已写的2个函数分别提取199it和贝恩的报告
    scrape_reports_199it(search_keywords, start_year, headers, worksheet, written_combinations, count)
    scrape_reports_bain(search_keywords, start_year, headers, worksheet, written_combinations, count)
    workbook.save(f"【{start_year}-至今】{'_'.join(search_keywords)}.xlsx")

# 主函数调用
if __name__ == "__main__":
    main()
    print("报告已经提取完成！")

# 记录代码结束的时间，计算代码运行时间
# end_time = time.time()
# run_time = end_time - start_time
# print(f"代码运行的时间为：{run_time}秒")


